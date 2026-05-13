"""Excel file validation checks"""

import logging
from typing import Dict, Any, List
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelValidator:
    """Validates Excel files in resources"""

    @staticmethod
    def check_sheet_count(file_path: str) -> Dict[str, Any]:
        """
        Check if Excel file has only 1 sheet
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Result dictionary with check status
        """
        result = {
            "check": "sheet_count",
            "file": Path(file_path).name,
            "status": "pass",
            "sheet_count": 0,
            "sheet_names": [],
            "issues": [],
        }

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            result["sheet_count"] = len(sheet_names)
            result["sheet_names"] = sheet_names

            if len(sheet_names) > 1:
                result["status"] = "fail"
                result["issues"].append(
                    f"Excel file has {len(sheet_names)} sheets. Expected 1 sheet."
                )

            workbook.close()
            
        except Exception as e:
            result["status"] = "fail"
            result["issues"].append(f"Error reading Excel file: {str(e)}")

        return result

    @staticmethod
    def check_pivot_tables(file_path: str) -> Dict[str, Any]:
        """
        Check if Excel file contains pivot tables
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Result dictionary with check status
        """
        result = {
            "check": "pivot_tables",
            "file": Path(file_path).name,
            "status": "pass",
            "has_pivot_tables": False,
            "pivot_count": 0,
            "issues": [],
        }

        try:
            workbook = openpyxl.load_workbook(file_path)
            
            pivot_count = 0
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                # Check if sheet has pivot tables
                if hasattr(ws, '_pivotTables') and ws._pivotTables:
                    pivot_count += len(ws._pivotTables)

            if pivot_count > 0:
                result["has_pivot_tables"] = True
                result["pivot_count"] = pivot_count
                result["status"] = "fail"
                result["issues"].append(
                    f"Excel file contains {pivot_count} pivot table(s). "
                    "Pivot tables should not be used for data storage."
                )

            workbook.close()
            
        except Exception as e:
            result["status"] = "warning"
            result["issues"].append(f"Could not check pivot tables: {str(e)}")

        return result

    @staticmethod
    def check_data_integrity(file_path: str) -> Dict[str, Any]:
        """
        Check basic data integrity in Excel file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Result dictionary with check status
        """
        result = {
            "check": "data_integrity",
            "file": Path(file_path).name,
            "status": "pass",
            "row_count": 0,
            "column_count": 0,
            "issues": [],
        }

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            ws = workbook.active
            
            # Get dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            
            result["row_count"] = max_row
            result["column_count"] = max_col

            if max_row == 0:
                result["status"] = "fail"
                result["issues"].append("Excel file has no data rows")

            if max_col == 0:
                result["status"] = "fail"
                result["issues"].append("Excel file has no data columns")

            # Check for empty cells in header row
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            empty_headers = sum(1 for cell in header_row if not cell)
            
            if empty_headers > 0:
                result["issues"].append(
                    f"Header row has {empty_headers} empty cells"
                )

            workbook.close()
            
        except Exception as e:
            result["status"] = "fail"
            result["issues"].append(f"Error reading Excel file: {str(e)}")

        return result

    @staticmethod
    def validate_excel_file(file_path: str) -> Dict[str, Any]:
        """
        Run all Excel validation checks on a file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Comprehensive validation result
        """
        checks = [
            ExcelValidator.check_sheet_count(file_path),
            ExcelValidator.check_pivot_tables(file_path),
            ExcelValidator.check_data_integrity(file_path),
        ]

        all_pass = all(check.get("status") == "pass" for check in checks)
        overall_status = "pass" if all_pass else (
            "fail" if any(check.get("status") == "fail" for check in checks)
            else "warning"
        )

        return {
            "excel_checks": checks,
            "overall_status": overall_status,
        }
